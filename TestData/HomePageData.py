import openpyxl


class HomePageData:
    # We have to wrap tuples, dictionaries etc. in list, that's the way pytest fixture accepts
    test_HomePage_data = [{"firstname": "Rahul",
                           "email": "rahul@gmail.com",
                           "password": "123456",
                           "gender": "Male"},
                          {"firstname": "Anshika",
                           "email": "anshika@gmail.com",
                           "password": "232323",
                           "gender": "Female"}]

    @staticmethod  # we don't need to create an object of the class to reach this method now
    def getTestData(test_case_name):  # we can remove self too after typing @staticmethod
        Dict = {}

        # This code loads the Excel workbook 'PythonDemo.xlsx' from the specified path using the openpyxl library.
        # The file is located on the user's desktop. The workbook can then be accessed for reading or modifying data.
        book = openpyxl.load_workbook("/Users/dogukanyavuz/Documents/PythonDemo.xlsx")

        # we got the active sheet
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
