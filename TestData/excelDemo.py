import openpyxl

# This code loads the Excel workbook 'PythonDemo.xlsx' from the specified path using the openpyxl library.
# The file is located on the user's desktop. The workbook can then be accessed for reading or modifying data.
book = openpyxl.load_workbook("/Users/dogukanyavuz/Documents/PythonDemo.xlsx")

# we got the active sheet
sheet = book.active

Dict = {}

cell = sheet.cell(row=1, column=2)
print(cell.value)  # firstname

# i write Dogukan into PythonDemo Excel file
sheet.cell(row=2, column=2).value = "Dogukan"
print(sheet.cell(row=2, column=2).value)  # Dogukan
print(sheet["A5"].value)  #Testcase4

print(sheet.max_row)
print(sheet.max_column)

